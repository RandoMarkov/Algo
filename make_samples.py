"""Make dated copies of the real template files so 'as-of' and 'evolution' are
demonstrable (the templates are a single date each).

Loads a template, overwrites the Synthese 'Date' cell, optionally scales a few
exposure cells to simulate drift, and saves to ./work_funds/.
"""
from pathlib import Path

import openpyxl

import config

BASE = Path(__file__).resolve().parent
TEMPLATES = {
    "equity": BASE / "incoming_samples/equity/EqtyTemplate.xlsx",
    "fixed_income": BASE / "incoming_samples/fixed_income/BondTemplate.xlsx",
    "alternatives": BASE / "incoming_samples/alternatives/AltTemplate.xlsx",
}
WORK = BASE / "work_funds"


def _set_date(ws, ddmmyyyy):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().lower() == "date":
            ws.cell(r, 2).value = ddmmyyyy
            return
    raise KeyError("Date label not found on Synthese")


def _scale_first_country(wb, factor):
    """Simulate exposure drift while PRESERVING the column sum: move weight
    between the first two geography buckets (as a real rebalance would)."""
    for name in ("GeographieDM", "Geographie"):
        if name in wb.sheetnames:
            ws = wb[name]
            a, b = ws.cell(2, 2), ws.cell(3, 2)
            if isinstance(a.value, (int, float)) and isinstance(b.value, (int, float)):
                delta = round(a.value * (factor - 1), 4)
                delta = min(delta, b.value)        # don't push the neighbour < 0
                a.value = round(a.value + delta, 4)
                b.value = round(b.value - delta, 4)
            return


def make_dated_copy(template_path, out_path, ddmmyyyy, drift=1.0):
    wb = openpyxl.load_workbook(template_path)          # keep literal values
    _set_date(wb[config.SHEET_SYNTHESIS], ddmmyyyy)
    if drift != 1.0:
        _scale_first_country(wb, drift)
    wb.save(out_path)
    return out_path


# (template key, ddmmyyyy, output filename, drift) for the demo
PLAN = [
    ("equity",       "31/03/2026", "EQ_2026-03-31.xlsx", 1.00),
    ("equity",       "18/06/2026", "EQ_2026-06-18.xlsx", 1.25),
    ("fixed_income", "31/03/2026", "FI_2026-03-31.xlsx", 1.00),
    ("fixed_income", "30/04/2026", "FI_2026-04-30.xlsx", 0.90),
    ("alternatives", "18/06/2026", "ALT_2026-06-18.xlsx", 1.00),
]


def main():
    WORK.mkdir(exist_ok=True)
    made = []
    for key, date, fname, drift in PLAN:
        p = make_dated_copy(TEMPLATES[key], WORK / fname, date, drift)
        made.append(p)
    print(f"Wrote {len(made)} dated fund files to {WORK}:")
    for p in made:
        print("  ", p.name)
    return made


if __name__ == "__main__":
    main()
