"""Read ONE fund workbook and record it into the database.

Handles all three fund types (Equity / Fixed Income / Alternatives): the type
is detected from the Synthese 'Categorie' cell, then the matching layout
profile in config.PROFILES drives which exposure sheets are read and how.

Pure Python + openpyxl: runs and is testable on any platform.
"""
from datetime import date, datetime
from pathlib import Path

import openpyxl

import config
import db


# --- helpers ----------------------------------------------------------------

def _norm_date(value) -> str:
    """Coerce a cell value to ISO 'YYYY-MM-DD'. Files use DD/MM/YYYY."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date value: {value!r}")


def _is_number(v) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace("%", "").replace(",", "."))
        return True
    except (TypeError, ValueError):
        return False


def _as_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).replace("%", "").replace(",", "."))


# --- Synthese: identity, numeric metrics, text attributes -------------------

def _load_synthese(ws):
    """Return (identity dict, metric rows, attribute rows).

    Walks every label/value pair in columns A/B. Known identity labels are
    pulled out; remaining numeric values become metrics, remaining text values
    become attributes.
    """
    identity, metrics, attributes = {}, [], []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        value = ws.cell(r, 2).value
        if label is None or str(label).strip() == "":
            continue
        key = str(label).strip()
        low = key.lower()
        if low in config.SYN_IDENTITY:
            identity[config.SYN_IDENTITY[low]] = value
        elif value is None or str(value).strip() == "":
            continue
        elif _is_number(value):
            metrics.append({"metric": key, "value": _as_float(value)})
        else:
            attributes.append({"attribute": key, "value_text": str(value).strip()})
    return identity, metrics, attributes


# --- generic exposure-sheet reader ------------------------------------------

def _load_exposure_sheet(ws, spec):
    """Yield {dimension, bucket, weight} rows for one exposure sheet spec."""
    last = {}  # forward-fill memory per label column
    for r in range(spec["start_row"], ws.max_row + 1):
        # build the bucket label from label_cols, forward-filling where asked
        parts = []
        for col in spec["label_cols"]:
            v = ws.cell(r, col).value
            if (v is None or str(v).strip() == "") and col in spec["ffill_cols"]:
                v = last.get(col)
            elif v is not None and str(v).strip() != "":
                last[col] = v
            if v is not None and str(v).strip() != "":
                parts.append(str(v).strip())
        label = " / ".join(parts)
        if not label or label.lower() in config.SKIP_LABELS:
            continue
        wv = ws.cell(r, spec["weight_col"]).value
        if not _is_number(wv):
            continue
        yield {"dimension": spec["dimension"], "bucket": label,
               "weight": _as_float(wv)}


def _load_holdings(ws):
    """Read the Inventaire sheet: locate header row, map columns, yield rows."""
    cols = config.HOLD_COLS
    wanted = set(cols.values())
    header_row, col_of = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        seen = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None}
        if wanted.issubset(seen.keys()):
            header_row = r
            col_of = {k: seen[h] for k, h in cols.items()}
            break
    if header_row is None:
        raise KeyError(f"Holdings headers {wanted} not found on {ws.title!r}")

    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        sec = ws.cell(r, col_of["security"]).value
        wv = ws.cell(r, col_of["weight"]).value
        if sec is None or not _is_number(wv):
            continue
        isin = ws.cell(r, col_of["isin"]).value
        out.append({"security": str(sec).strip(),
                    "isin": "" if isin is None else str(isin).strip(),
                    "weight": _as_float(wv)})
    return out


# --- top level --------------------------------------------------------------

def ingest_fund_file(file_path: str | Path, conn=None) -> dict:
    """Record one fund workbook (any of the 3 types) into the DB.

    Type is auto-detected from Synthese 'Categorie'. Idempotent: re-recording
    the same (fund, date) replaces that snapshot. Returns a summary dict.
    """
    file_path = Path(file_path)
    owns = conn is None
    conn = conn or db.connect()
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        identity, metrics, attributes = _load_synthese(wb[config.SHEET_SYNTHESIS])
        fund_id = str(identity["fund_id"]).strip()
        name = str(identity.get("name", "")).strip()
        category = str(identity.get("category", "")).strip()
        as_of = _norm_date(identity["as_of"])
        profile_name = config.profile_for_category(category)
        profile = config.PROFILES[profile_name]

        # Several sheets can feed one dimension (e.g. Devise DM+EM -> currency),
        # and a bucket can repeat across or within those sheets (USD appears in
        # both DM and EM). Aggregate by (dimension, bucket), summing weights.
        agg = {}
        sheets_read = []
        for spec in profile["exposures"]:
            if spec["sheet"] not in wb.sheetnames:
                continue  # tolerate a missing optional sheet
            for e in _load_exposure_sheet(wb[spec["sheet"]], spec):
                key = (e["dimension"], e["bucket"])
                agg[key] = agg.get(key, 0.0) + e["weight"]
            sheets_read.append(spec["sheet"])
        exposures = [{"dimension": d, "bucket": b, "weight": w}
                     for (d, b), w in agg.items()]

        holdings = _load_holdings(wb[config.HOLDINGS_SHEET])
        wb.close()

        # write
        db.upsert_fund(conn, fund_id, name, None)
        db.replace_snapshot(conn, "fund_metrics", fund_id, as_of,
                            [{"fund_id": fund_id, "as_of_date": as_of, **m}
                             for m in metrics])
        db.replace_snapshot(conn, "fund_attributes", fund_id, as_of,
                            [{"fund_id": fund_id, "as_of_date": as_of, **a}
                             for a in attributes]
                            + [{"fund_id": fund_id, "as_of_date": as_of,
                                "attribute": "Profile", "value_text": profile_name}])
        db.replace_snapshot(conn, "fund_exposures", fund_id, as_of,
                            [{"fund_id": fund_id, "as_of_date": as_of, **e}
                             for e in exposures])
        db.replace_snapshot(conn, "fund_holdings", fund_id, as_of,
                            [{"fund_id": fund_id, "as_of_date": as_of, **h}
                             for h in holdings])
        conn.commit()
    finally:
        if owns:
            conn.close()

    dims = sorted({e["dimension"] for e in exposures})
    return {"fund_id": fund_id, "name": name, "category": category,
            "profile": profile_name, "as_of_date": as_of,
            "n_holdings": len(holdings), "n_exposure_rows": len(exposures),
            "dimensions": dims, "n_metrics": len(metrics),
            "sheets_read": sheets_read}
