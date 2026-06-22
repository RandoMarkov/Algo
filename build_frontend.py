"""Generate the front-end workbook (PortfolioTool.xlsx) with openpyxl.

openpyxl cannot create VBA or buttons (that must happen in Excel on Windows),
but it builds everything else: the Control inputs, the Portfolio allocation
table, a pre-filled xlwings.conf, and an Instructions sheet. After running this
you import FundTool.bas, draw 3 buttons, and Save As .xlsm — see SETUP.md.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

import config

BASE = Path(__file__).resolve().parent
OUT = BASE / "PortfolioTool.xlsx"

# --- styling helpers --------------------------------------------------------
TITLE = Font(size=16, bold=True, color="FFFFFF")
HDR = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
NOTE = Font(italic=True, color="555555")
DARK = PatternFill("solid", fgColor="1F4E78")
MID = PatternFill("solid", fgColor="2E75B6")
INPUT = PatternFill("solid", fgColor="FFF2CC")   # yellow = "type here"
GREY = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell(ws, ref, value=None, font=None, fill=None, align=None, border=None):
    c = ws[ref]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


# --- Control sheet ----------------------------------------------------------
def build_control(wb):
    ws = wb.active
    ws.title = config.CTRL_SHEET
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46

    ws.merge_cells("A1:B1")
    _cell(ws, "A1", "Fund & Portfolio Tool", TITLE, DARK,
          Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 28

    # inputs (cells referenced by config.CTRL_CELLS)
    _cell(ws, "A2", "Fund file to import", BOLD)
    _cell(ws, "B2", "EQ_2026-06-18.xlsx", None, INPUT, None, BOX)
    _cell(ws, "A3", "Portfolio name", BOLD)
    _cell(ws, "B3", "MULTI", None, INPUT, None, BOX)
    _cell(ws, "A4", "As-of date (YYYY-MM-DD)", BOLD)
    _cell(ws, "B4", "2026-06-20", None, INPUT, None, BOX)
    # evolution inputs (button ④)
    _cell(ws, "A5", "Evolution dimension", BOLD)
    _cell(ws, "B5", "geography_country", None, INPUT, None, BOX)
    _cell(ws, "A6", "Evolution start (YYYY-MM-DD)", BOLD)
    _cell(ws, "B6", "2026-01-01", None, INPUT, None, BOX)
    _cell(ws, "A7", "Evolution end (YYYY-MM-DD)", BOLD)
    _cell(ws, "B7", "2026-06-20", None, INPUT, None, BOX)
    _cell(ws, "A8", "Evolution step (months)", BOLD)
    _cell(ws, "B8", 1, None, INPUT, None, BOX)
    for r in range(2, 9):
        ws.cell(r, 1).alignment = Alignment(vertical="center")

    # dropdown of valid dimension keys for B5
    dims = ",".join(config.DIM_LABELS.keys())
    dv = DataValidation(type="list", formula1=f'"{dims}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B5")

    _cell(ws, "A10", "Status", BOLD)
    _cell(ws, "B10", "(messages from the buttons appear here)", NOTE, GREY, None, BOX)

    # button legend
    _cell(ws, "A12", "Buttons", HDR, MID)
    _cell(ws, "B12", "Macro to assign (see SETUP.md)", HDR, MID)
    rows = [("① Record fund", "Btn_RecordFund  — import the file in B2 into the DB"),
            ("② Compute portfolio", "Btn_Compute  — look-through for B3 as-of B4 → Results"),
            ("③ Make charts", "Btn_Charts  — draw charts from Results → Charts"),
            ("④ Evolution", "Btn_Evolution  — B5 over B6→B7 (step B8) → Evolution")]
    for i, (a, b) in enumerate(rows, start=13):
        _cell(ws, f"A{i}", a, BOLD, None, None, BOX)
        _cell(ws, f"B{i}", b, None, None, Alignment(vertical="center"), BOX)
        ws.row_dimensions[i].height = 20

    _cell(ws, "A18", "Workflow", HDR, MID)
    ws.merge_cells("A19:B19")
    _cell(ws, "A19",
          "1) Fund file name in B2 → click ①  (repeat per fund).  "
          "2) Fill the Portfolio sheet, set B3/B4 → click ② then ③.  "
          "3) For a trend, set B5–B8 → click ④.",
          NOTE, GREY, Alignment(wrap_text=True, vertical="top"))
    ws.merge_cells("A20:B20")
    ws.row_dimensions[19].height = 44


# --- Portfolio sheet --------------------------------------------------------
def build_portfolio(wb):
    ws = wb.create_sheet(config.PORTFOLIO_SHEET)
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 18), ("B", 18), ("C", 12)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:C1")
    _cell(ws, "A1", "Portfolio allocations  (time-varying — add a new dated row to rebalance)",
          NOTE)
    for col, name in (("A", "Fund ID"), ("B", "Effective Date"), ("C", "Weight")):
        _cell(ws, f"{col}2", name, HDR, MID, Alignment(horizontal="center"), BOX)

    # sample rows: the three template ISINs, two rebalance dates
    sample = [("LU1670707527", "2026-01-01", 0.50),
              ("LU1333337",    "2026-01-01", 0.50),
              ("LU1670707527", "2026-06-01", 0.40),
              ("LU1333337",    "2026-06-01", 0.35),
              ("LU70707527ALT", "2026-06-01", 0.25)]
    for i, (fid, eff, w) in enumerate(sample, start=3):
        _cell(ws, f"A{i}", fid, None, None, None, BOX)
        _cell(ws, f"B{i}", eff, None, None, None, BOX)
        _cell(ws, f"C{i}", w, None, None, Alignment(horizontal="center"), BOX)

    # weights must be numeric 0..1
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1",
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Weight must be between 0 and 1."
    ws.add_data_validation(dv)
    dv.add("C3:C200")


# --- empty Results / Charts placeholders ------------------------------------
def build_placeholders(wb):
    for name, note in ((config.RESULTS_SHEET, "Computed numbers appear here after ② Compute."),
                       (config.CHARTS_SHEET, "Charts appear here after ③ Make charts."),
                       (config.EVOLUTION_SHEET, "Time-series + line chart appear here after ④ Evolution.")):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        _cell(ws, "A1", note, NOTE)


# --- xlwings.conf (pre-configures the Python side) --------------------------
def build_conf(wb):
    ws = wb.create_sheet("xlwings.conf")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    _cell(ws, "A1", "xlwings configuration — EDIT the two paths for your PC", NOTE)
    rows = [
        ("Interpreter_Win", r"%(folder)s\.venv\Scripts\pythonw.exe"),
        ("PYTHONPATH", r"%(folder)s"),
        ("Show Console", "False"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        _cell(ws, f"A{i}", k, BOLD)
        _cell(ws, f"B{i}", v)
    _cell(ws, "A6",
          "%(folder)s = the folder this workbook is saved in. If you used a "
          "different venv, point Interpreter_Win at its pythonw.exe.", NOTE)
    ws.merge_cells("A6:B7")
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = Workbook()
    build_control(wb)
    build_portfolio(wb)
    build_placeholders(wb)
    build_conf(wb)
    wb.active = 0
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
