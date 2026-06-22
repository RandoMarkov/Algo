"""Render a computed result to a standalone .xlsx with native Excel charts.

Headless counterpart of the third button: numbers on a Results sheet, pie/bar
charts on a Charts sheet, openable in Excel with no macros. Dimensions are
whatever the funds reported, so this adapts to all three fund types.
"""
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference

import config

PIE_MAX_SLICES = 8     # above this, use a bar chart instead
TOP_HOLDINGS = 15


def _write_table(ws, start_row, title, df):
    ws.cell(start_row, 1, title).font = ws.cell(start_row, 1).font.copy(bold=True)
    hdr = start_row + 1
    ws.cell(hdr, 1, df.columns[0])
    ws.cell(hdr, 2, df.columns[1])
    r = hdr
    for _, row in df.iterrows():
        r += 1
        ws.cell(r, 1, row.iloc[0])
        ws.cell(r, 2, float(row.iloc[1]))
    return hdr, hdr + 1, r          # header row, first data, last data


def render_results(result: dict, out_path: str):
    wb = Workbook()
    res = wb.active
    res.title = config.RESULTS_SHEET

    res["A1"], res["B1"] = "Portfolio", result["portfolio_id"]
    res["A2"], res["B2"] = "As-of date", result["as_of_date"]
    res["A3"], res["B3"] = "Invested weight", round(result["invested_weight"], 6)
    res["A4"], res["B4"] = "Cash residual", round(result["cash_residual"], 6)

    ranges = {}        # title -> (hdr,first,last, chart_kind)
    row = 6

    # one section per exposure dimension (sorted, friendly labels)
    for dim in sorted(result["exposures"]):
        df = result["exposures"][dim]
        df = df[df["weight"].abs() > 1e-9]                 # drop empty buckets
        if df.empty:
            continue
        label = config.DIM_LABELS.get(dim, dim)
        cov = result["coverage"].get(dim, 0.0)
        title = f"{label}  (coverage {cov:.0%})"
        hdr, first, last = _write_table(res, row, title, df)
        kind = "pie" if len(df) <= PIE_MAX_SLICES else "bar"
        ranges[title] = (hdr, first, last, kind)
        row = last + 2

    # holdings (look-through), top N -> bar
    hd = result["holdings"]
    hd = hd[hd["weight"].abs() > 1e-9].head(TOP_HOLDINGS)
    if not hd.empty:
        hdr, first, last = _write_table(res, row, "Top holdings (look-through)", hd)
        ranges["Top holdings (look-through)"] = (hdr, first, last, "bar")
        row = last + 2

    # metrics table (no chart)
    if not result["metrics"].empty:
        _write_table(res, row, "Weighted metrics", result["metrics"])

    # --- charts sheet ---
    ch = wb.create_sheet(config.CHARTS_SHEET)
    anchor = 1
    for title, (hdr, first, last, kind) in ranges.items():
        chart = PieChart() if kind == "pie" else BarChart()
        if kind == "bar":
            chart.type = "bar"
        chart.title = title
        labels = Reference(res, min_col=1, min_row=first, max_row=last)
        data = Reference(res, min_col=2, min_row=hdr, max_row=last)   # incl header
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height, chart.width = 8, 16
        ch.add_chart(chart, f"A{anchor}")
        anchor += 16

    wb.save(out_path)
    return out_path


def render_evolution(evo, out_path: str, title: str = "Exposure evolution"):
    """Write a time-series table (rows=date, cols=bucket) + a line chart."""
    wb = Workbook()
    ws = wb.active
    ws.title = config.EVOLUTION_SHEET

    ws.cell(1, 1, "date")
    for j, col in enumerate(evo.columns, start=2):
        ws.cell(1, j, str(col))
    for i, (idx, row) in enumerate(evo.iterrows(), start=2):
        ws.cell(i, 1, str(idx))
        for j, col in enumerate(evo.columns, start=2):
            ws.cell(i, j, float(row[col]))

    nrow, ncol = 1 + len(evo), 1 + evo.shape[1]
    data = Reference(ws, min_col=2, max_col=ncol, min_row=1, max_row=nrow)
    cats = Reference(ws, min_col=1, min_row=2, max_row=nrow)

    # line chart (per-bucket trend) and stacked-area chart (composition) side by side
    line = LineChart()
    line.title = f"{title} — lines"
    line.height, line.width = 10, 22
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, f"A{nrow + 2}")

    area = AreaChart()
    area.grouping = "stacked"
    area.overlap = 100
    area.title = f"{title} — stacked area"
    area.height, area.width = 10, 22
    area.add_data(data, titles_from_data=True)
    area.set_categories(cats)
    ws.add_chart(area, f"M{nrow + 2}")

    wb.save(out_path)
    return out_path
